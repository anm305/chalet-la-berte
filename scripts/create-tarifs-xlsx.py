#!/usr/bin/env python3
"""Create / update the Excel tariff workbook and expand it to data/rates.csv."""

from __future__ import annotations

import csv
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
XLSX = DATA / "tarifs.xlsx"
CSV_OUT = DATA / "rates.csv"

HEADER_FILL = PatternFill("solid", fgColor="12211B")
HEADER_FONT = Font(color="FAFAF6", bold=True)
HINT_FILL = PatternFill("solid", fgColor="E4E6DE")
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def parse_excel_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def parse_price(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    n = float(str(value).replace("CHF", "").replace(" ", "").replace(",", "."))
    return n if n > 0 else None


def style_header(ws, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def create_template():
    DATA.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Instructions ---
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Chalet La Berte — Tarifs site web"
    ws["A1"].font = Font(size=16, bold=True, color="12211B")
    lines = [
        "",
        "Ce fichier est la source des prix affichés sur le site (avant remise directe −10 %).",
        "",
        "1. Onglet SAISONS : définis des plages de dates + prix Airbnb / nuit (CHF).",
        "2. Onglet EXCEPTIONS : prix pour des dates précises (prioritaires sur les saisons).",
        "3. Enregistre le fichier.",
        "4. Dans le terminal du projet :",
        "      npm run import:rates",
        "5. Puis publie :",
        '      git add data/rates.csv data/tarifs.xlsx && git commit -m "Update rates" && git push',
        "",
        "Le site applique ensuite PUBLIC_DIRECT_DISCOUNT_PERCENT (défaut 10 %).",
        "Exemple : Airbnb 800 CHF → site 720 CHF.",
        "",
        "Astuce : commence par adapter les saisons d’exemple ci-dessous à TES vrais tarifs Airbnb.",
    ]
    for i, line in enumerate(lines, start=2):
        ws[f"A{i}"] = line
    ws.column_dimensions["A"].width = 92

    # --- Saisons ---
    ws = wb.create_sheet("Saisons")
    ws.append(["debut", "fin", "prix_airbnb_chf", "label"])
    style_header(ws, 4)
    # Example seasons spanning ~18 months from Aug 2026
    examples = [
        (date(2026, 4, 1), date(2026, 6, 30), 750, "Basse saison printemps"),
        (date(2026, 7, 1), date(2026, 8, 31), 950, "Haute saison été"),
        (date(2026, 9, 1), date(2026, 11, 30), 780, "Automne"),
        (date(2026, 12, 1), date(2026, 12, 19), 850, "Hiver"),
        (date(2026, 12, 20), date(2027, 1, 5), 1200, "Noël / Nouvel An"),
        (date(2027, 1, 6), date(2027, 3, 31), 880, "Hiver ski"),
        (date(2027, 4, 1), date(2027, 6, 30), 750, "Basse saison printemps"),
        (date(2027, 7, 1), date(2027, 8, 31), 950, "Haute saison été"),
    ]
    for row in examples:
        ws.append(list(row))
    for col, width in enumerate([14, 14, 16, 28], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN
            if cell.column <= 2 and isinstance(cell.value, date):
                cell.number_format = "YYYY-MM-DD"

    note = ws.cell(ws.max_row + 2, 1, "⚠️ Remplace les prix d’exemple par tes tarifs Airbnb réels.")
    note.fill = HINT_FILL

    # --- Exceptions ---
    ws = wb.create_sheet("Exceptions")
    ws.append(["date", "prix_airbnb_chf", "note"])
    style_header(ws, 3)
    ws.append([date(2026, 8, 1), 1100, "Exemple : week-end férié (à modifier ou supprimer)"])
    for col, width in enumerate([14, 16, 40], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws["A2"].number_format = "YYYY-MM-DD"

    wb.save(XLSX)
    print(f"Modèle créé : {XLSX}")


def expand_rates(wb):
    """Expand Saisons + Exceptions → list of (date, price)."""
    by_date: dict[date, float] = {}

    if "Saisons" in wb.sheetnames:
        ws = wb["Saisons"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            start = parse_excel_date(row[0])
            end = parse_excel_date(row[1])
            price = parse_price(row[2] if len(row) > 2 else None)
            if not start or not end or price is None:
                continue
            if end < start:
                start, end = end, start
            cur = start
            while cur <= end:
                by_date[cur] = price
                cur += timedelta(days=1)

    if "Exceptions" in wb.sheetnames:
        ws = wb["Exceptions"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            d = parse_excel_date(row[0])
            price = parse_price(row[1] if len(row) > 1 else None)
            if d and price is not None:
                by_date[d] = price

    # Optional sheet Tarifs (day-by-day overrides / full list)
    if "Tarifs" in wb.sheetnames:
        ws = wb["Tarifs"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            d = parse_excel_date(row[0])
            price = parse_price(row[1] if len(row) > 1 else None)
            if d and price is not None:
                by_date[d] = price

    return sorted(by_date.items(), key=lambda x: x[0])


def import_workbook(path: Path | None = None):
    path = path or XLSX
    if not path.exists():
        print(f"Fichier introuvable : {path}")
        print("Lance d’abord : python3 scripts/create-tarifs-xlsx.py")
        sys.exit(1)

    wb = load_workbook(path, data_only=True)
    rows = expand_rates(wb)
    if not rows:
        print("Aucun tarif trouvé. Remplis l’onglet Saisons (debut, fin, prix_airbnb_chf).")
        sys.exit(1)

    DATA.mkdir(parents=True, exist_ok=True)
    with CSV_OUT.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["# Généré depuis " + path.name])
        w.writerow(["# " + datetime.now().isoformat(timespec="seconds")])
        w.writerow(["date", "price"])
        for d, price in rows:
            # store as int if whole number
            p = int(price) if float(price).is_integer() else round(price, 2)
            w.writerow([d.isoformat(), p])

    print(f"OK — {len(rows)} nuits → {CSV_OUT}")
    print(f"Période : {rows[0][0]} → {rows[-1][0]}")
    print("Publie avec : git add data/ && git commit -m \"Update rates\" && git push")


def main():
    args = sys.argv[1:]
    if not args or args[0] in ("init", "create", "--init"):
        create_template()
        if not args or args[0] in ("init", "create", "--init"):
            # also import example so rates.csv is ready for testing
            if "--no-import" not in args:
                import_workbook(XLSX)
        return

    if args[0] in ("import", "--import"):
        src = Path(args[1]).expanduser() if len(args) > 1 else XLSX
        import_workbook(src)
        return

    # path given directly
    src = Path(args[0]).expanduser()
    if src.suffix.lower() in (".xlsx", ".xlsm"):
        import_workbook(src)
    else:
        print("Usage:")
        print("  python3 scripts/create-tarifs-xlsx.py init")
        print("  python3 scripts/create-tarifs-xlsx.py import [fichier.xlsx]")
        sys.exit(1)


if __name__ == "__main__":
    main()
